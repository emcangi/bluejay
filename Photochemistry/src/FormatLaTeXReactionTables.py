import pandas as pd
import copy
import re
from openpyxl import load_workbook

def format_spc(s):
    """
    Replaces certain sequences in a species name string, s, with LaTeX formatting.
    """
    s = s.replace("pl", "^+")
    s = s.replace("E", "e^-")
    s = s.replace("2", "_2")
    s = s.replace("3", "_3")
    s = s.replace("up_2D", "(^2D)")
    s = s.replace("1D", "(^1D)")
    return s
    
def participant_str(r1, r2="", r3=""):
    """
    Constructs a string that prints the reactants of a chemical reaction.

    r1: Reactant 1
    r2: Reactant 2 (optional)
    r3: Reactant 3 (optional)
    """
    if r1 != "":
        if (r2 != "") & (r2 != "none"):
            if (r3 != "") & (r3 != "none"):
                return f"{format_spc(r1)} + {format_spc(r2)} + {format_spc(r3)}"
            else:
                return f"{format_spc(r1)} + {format_spc(r2)}"
        else:
            return f"{format_spc(r1)}"
    else:
        return ""

def br_str(br):
    # """
    # Fills a string with the branching ratio, if present.
    # """
    # if (br == "") | (int(float(br)) == 1):
    #     return "" 
    """Fills a string with the branching ratio, if present."""
    # Avoid evaluating the second condition when ``br`` is empty. Using
    # ``or`` prevents ``int(float(br))`` from raising ``ValueError``.
    if br == "" or int(float(br)) == 1:
        return ""
    else:
        return f"{br}"
    
def f_str(f):
    """
    Constructs the string containing the Troe parameter, f.
    """
    if f == "":
        return "" 
    else:
        if float(f) != 0:
            return f"{f}"
        else: 
            return ""
   
def mscale_str(m2, m1, p):
    """
    Constructs the string used to represent mass scaling, applied to 
    certain H-bearing reactions to produce a deuterated analogue reaction.

    m2: Mass of the heavier isotope
    m1: Mass of the lighter isotope
    p: power to which to raise the ratio of masses
    """
    if m2 != 1:
        if p == -0.5:
            return f"$\sqrt{{ \\frac{{{m1}}}{{{m2}}} }}$"
        else:
            return f"$\left(\\frac{{{m2}}}{{{m1}}}\right)^{{{p}}}$"
    else:
        return ""
    
# Enter in the rate coefficient column
def a_str(a):
    """
    Constructs the string containing A (leading factor) which is part of the rate coefficient.

    a: the leading factor for k_inf definitions.
    """
    Astring = str("{:.2e}".format(a))
    if "e-" in Astring:
        Astring = Astring.replace("e-0", r" \times 10^{-")
        Astring = Astring.replace("e-", r" \times 10^{-")
        Astring = Astring + "}"

    return Astring

def b_str(b, T="T_i"):
    """
    Constructs the string containing B (power on T) which is part of the rate coefficient.

    Note that this does not include a factor of 1/300, which is included in some reaction rate coefficient
    formats, but not in the network that Roger supplied.

    b: the power on T for k_inf definitions.
    T: a string to fill in for temperature. e.g. Tn, Ti, Te for neutrals, ions, elecrons.
    """
    if b != 0:
        return f"\left({T}\right)^{{{b}}}"#f"\left(\\frac{{{T}}}{{300}}\right)^{{{b}}}"
    else:
        return ""

def c_str(c, T="T_i"):
    """
    Constructs the string containing C (exponential factor) which is part of the rate coefficient.

    Note that it is assumed that c contains the sign (positive or negative).

    c: the exponential factor for k_inf definitions.
    """
    if c != 0:
        return f"e^{{{c:g}/{T}}}" # :g removes any trailing zeros after the decimal, but retains non-zero decimals.
    else:
        return ""
    
def k_str(a, b, c, a0="", b0="", c0="", T="T_i"):
    """
    Constructs the rate coefficient k using A, B, and C.

    a0, b0, and c0 refer to factors to construct k0. 

    if k0 factors are present, this will construct a multi-line cell with k_inf, k_0, and a comment to see the text
    for further explanation.
    """
    if (a0 != 0) & (a0 != ""):
        return r"\makecell[l]{See text \\ k$_{\infty}="+f"{a_str(a)}{b_str(b, T=T)}{c_str(c, T=T)}$ \\\ k$_0={a_str(a0)}{b_str(b0, T=T)}{c_str(c0, T=T)}$" + r"}"
    else:
        return f"${a_str(a)}{b_str(b, T=T)}{c_str(c, T=T)}$"

def typestr(t):
    """
    Constructs a string which prints the reaction type, t, as according to Vuitton 2019.
    """
    if t != "":
        if int(t) > 2:
            return str(int(t))
        else:
            return ""
    else:
        return ""
    
def refstr(r):
    """
    This needs to be made a lot more efficient/elegant. Like probably loop through a dictionary based on an ordered list of the keys,
    or something that searches more efficiently.

    This takes a given string and replaces some text with other text. It's used to shorten the reference tags so that the reference
    column in the main reaction rate table can be small.
    """
    r = r.replace("Mass scaling???", "Est.")
    r = r.replace("Mass scaling", "Est.")
    r = r.replace("estimated", "Est.")
    r = r.replace("estimate", "Est.")
       
    r = r.replace("Anicich2003", "A03")
    r = r.replace("Anicich93, Klippenstein2010", "A93,Kl10")
    r = r.replace("Anicich1993", "A93")
    r = r.replace("Anicich93", "A93")
    
    r = r.replace("Atkinson89", "At89")
    r = r.replace("Antipov09", "Ant09")
    
    r = r.replace("Baulch05 (H analogue rate)", "Est.")
    r = r.replace("Baulch05, Horowitz78", "B05")
    r = r.replace("Baulch05", "B05")
    r = r.replace("Baulch92", "B92")
    r = r.replace("Borodi2009", "Bo09")
    r = r.replace("Burkholder2020.", "Bu20")
    r = r.replace("Burkholder2020", "Bu20")
    r = r.replace("Burkholder2020, Chaffin2017", "Bu20")
    r = r.replace("Brune83", "Br83")
    
    r = r.replace("Campbell73", "Cam73")
    r = r.replace("Cazaux2010", "C10")
    r = r.replace("Chaffin2017", "Est.")
    
    r = r.replace("Deighan2012", "D12")
    r = r.replace("Deighan12", "D12")
    r = r.replace("Davidson78", "Da78")
    r = r.replace("Doroshenko92", "Do92")
    r = r.replace("Dutuit13", "Du13")
    
    r = r.replace("Fehsenfeld1970", "Feh70")
    r = r.replace("Fletcher76", "Fl76")
    r = r.replace("Fox1993", "F93")
    r = r.replace("estimated (Fox2015)", "F15")
    r = r.replace("Fox2015, Vejby-Christensen1998, Hellberg2003", "F15")
    r = r.replace("Fox2015", "F15")
    
    r = r.replace("rate Korolov2009, BR Geppert05", "GK")
    r = r.replace("Geppert04,Herd90", "G04,H90")
    r = r.replace("Geppert05", "G05")
    r = r.replace("Geppert00", "G00")
    
    r = r.replace("Gustafsson15", "Gu15")
    r = r.replace("Heidner73", "Hei73")
    r = r.replace("Herron99", "Her99")
    r = r.replace("Hierl97", "Hi97")
    
    r = r.replace("Jensen00", "J00")
    
    r = r.replace("KIDA (Baulch2005)", "KIDA")
    
    r = r.replace("Korolov2009", "K09")
    r = r.replace("LeGarrec03", "L03")
    r = r.replace("Lindinger1975", "UMIST")
    r = r.replace("Milligan00,Klippenstein10", "M00,Kl10")
    r = r.replace("Mitchell05", "Mi05")
    r = r.replace("Mitchell90", "Mi90")
    
    r = r.replace("Nahar97", "N97")
    r = r.replace("NIST (Shavitt 1959)", "NIST")
    r = r.replace("NIST (Mitchell & Le Roy 1973)", "NIST")
    r = r.replace("NIST apparently", "NIST")
      
    r = r.replace("Okabe78", "O78")
    
    r = r.replace("Paraskevopoulos71", "P71")
    r = r.replace("Praxmarer94", "Pr94")
    
    r = r.replace("Rosen98", "R98")
    r = r.replace("Rosen00", "R00")
    
    r = r.replace("Sander15, Burkholder2020", "Bu20")
    r = r.replace("Sander15", "S15")
    r = r.replace("Sander2011", "S11")
    r = r.replace(" Sander2011, NIST", "S11")
    r = r.replace("Scott97", "Sc97")
    r = r.replace("Scott98", "Sc98")
    r = r.replace("Schunk09", "SN09")
    r = r.replace("Smith02", "Sm02")
    r = r.replace("Spanel93", "Sp93")
    
    r = r.replace("Tully75", "Tu75")
    r = r.replace("Tsang86", "Ts86")
    r = r.replace("Tsang91", "Ts91")
    
    r = r.replace("UMIST RATE12", "UMIST")
    
    r = r.replace("Viggiano80", "V80")
    r = r.replace("Viggiano05", "V05")

    r = r.replace("Yung1988, Yung1989, NIST", "NIST")
    r = r.replace("Yung1988: 0.71xH-analogue", "Y88")
    r = r.replace("Yung1988", "Y88")
    r = r.replace("Yung 1988", "Y88")
    r = r.replace("Yung1989", "Y89")
    
    r = r.replace("Zanchet09", "Z09")

    return r

def assign_reaction_numbers(df, opt_start=1):
    """
    Given a datarame of reaction parameters based on the reaction network spreadsheet, this function will assign reaction numbers
    including subnumbers a, b, c... etc for multiple product channels.

    df: dataframe containing entries from the reaction spreadsheet.
    opt_start: optional number at which to start the reaction numbering. This can be used if working with more than one logical 
               group of reaction types. e.g. reactions 1-14 might be photodissociation, but then separately you process 15-90 which are
               neutral, bimolecular reactions.
    """
    
    rxn_num = opt_start

    # dictionary to hold compared reactants
    comp = {"R1": "", "R2": "", "R3": ""}

    di=-2
        
    # Identify which columns are present in this dataframe (by name)
    rcols = set(comp.keys()).intersection(df.columns) 
    
    for index, row in df.iterrows():
        L = "a" # the reaction numbre extension

        # print(f"Comparing Rxn {rxn_num}: {df.loc[index, 'R1']} and {df.loc[index, 'R2']} to {comparison_r1} and {comparison_r2}")

        reactants_this_row = [df.loc[index, r] for r in rcols]
        this_row_str = "+".join(reactants_this_row)

        # We need to compare the current reaction and the next to see if it's the same reactants with different products to help
        # determine if we need to apply a, b, c, etc.
        # Note that this starts out on the first row with the 'comp' dictionary having empty entries of course.
        to_compare = [comp[r] for r in rcols] 
        join_compare = "+".join(to_compare)
        
        # print(f"comparing {to_compare} and {reactants_this_row}")

        # This block runs when it is determined that the comparison reaction and the current reaction are different,
        # i.e. it is time to assign a new reaction number.
        if to_compare != reactants_this_row: 
            # print(f"row {index} not a match: {join_compare} != {this_row_str}")
            df.loc[index, "Num"] = rxn_num # give it a number 

            # Set this reaction as the new reaction for comparing with the next in the dataframe
            for r in rcols:
                comp[r] = df.loc[index, r]

            # print(f"New comparisons {comparison_r1} and {comparison_r2}")

            # prepare to move on to the next
            rxn_num += 1
            comp_ind = index
            L = "a"
        else: # This block runs if this reaction and the comparison have the same reactants, i.e. different product branches of the same reaction:
            # print(f"row {index} match: {join_compare} = {this_row_str}")
            branchno = index - comp_ind + 1
            # print(f"Branch number is {branchno}")
            this_num = rxn_num - 1
            # print(f"Reaction number to use is {this_num}")

            # Reletter the number of the first in the set with an a
            df.loc[comp_ind, "Num"] = str(this_num) + L 

            # Give this one a b, c, or whatever is appropriate
            df.loc[index, "Num"] = str(this_num) + chr(ord(L) + branchno - 1)

            # Clear out the reactants because we don't need them printed again
            for r in rcols:
                df.loc[index, r] = ""
                
            L = chr(ord(L)+1) # increment the sub-reaction letter in alphabetical order.

    return rxn_num # will return the NEXT reaction number -- so we can know what to start with on the next batch.
    


# ===================================================================================================#
#                                       Script Start                                                 #
# ===================================================================================================#

# Dont cut off long cells!!!
pd.set_option('max_colwidth', None)

# Just a folder from which to get a parameter file to be able to list all_species.
main_folder = input("Enter the full pathname to the folder in which the reaction spreadsheet exists: ")
# print(f"Using the following file to extract the all species list: {main_folder}\n")
filename = input("Enter the filename with extension: ")

which = input("Do you want to do D reactions only (D only), non-D reactions (no D) or all (just press enter)?")

whichsort = input("Do you want to sort by reactants (R/r) or column rate (C/c)? ")
sort_by_reactants = True if whichsort in ["R", "r"] else False

# We read the excel file as a dataframe
df_photod = pd.read_excel(f"{main_folder}{filename}", sheet_name="Photodissociation", 
                    index_col=None, usecols=["R1", "P1", "P2", "P3", "BR", "Reference", "ColumnRate"])

df_photoi = pd.read_excel(f"{main_folder}{filename}", sheet_name="Photoionization", 
                    index_col=None, usecols=["R1", "P1", "P2", "BR", "Reference", "ColumnRate"])

df_neutrals = pd.read_excel(f"{main_folder}{filename}", sheet_name="Neutral reactions", 
                    index_col=None, usecols=["R1", "R2", "R3", "P1", "P2", "P3", "type", "M2", "M1", "pow", "BR", "kA", "kB", "kC", "k0A", "k0B", "k0C", "F", "Reference", "ColumnRate"])

df_ions = pd.read_excel(f"{main_folder}{filename}", sheet_name="Ion reactions", 
                    index_col=None, usecols=["R1", "R2", "P1", "P2", "P3", "M2", "M1", "pow", "BR", "kA", "kB", "kC", "k0A", "k0B", "k0C", "Reference", "ColumnRate"])

# Needful things

all_species = []
wb = load_workbook(filename = main_folder + "PARAMETERS.xlsx")
ws = wb["SpeciesLists"]
for col in ws.iter_cols(min_row=2, max_col=1, max_row=100):
    for cell in col:
        if cell.value != None:
            all_species.append(cell.value)
            
D_species = [a  for a in all_species if (("D" in a) & (a not in ["O1D", "Nup2D"]))]
non_D_species = list(sorted(set(all_species).difference(set(D_species))))

# Designate lists 
if which=="D only":
    df_photod_D = df_photod[ df_photod["R1"].isin(D_species) ]
    df_photoi_D = df_photoi[ df_photoi["R1"].isin(D_species) ]

    photo_df = pd.concat([df_photod_D, df_photoi_D], ignore_index=True)
    n_df =  df_neutrals[ df_neutrals["R1"].isin(D_species) | df_neutrals["R2"].isin(D_species) | df_neutrals["R3"].isin(D_species) ]
    i_df =  df_ions[ df_ions["R1"].isin(D_species) | df_ions["R2"].isin(D_species) ]
elif which=="no D":
    df_photod_noD = df_photod[ ~df_photod["R1"].isin(D_species) ]
    df_photoi_noD = df_photoi[ ~df_photoi["R1"].isin(D_species) ]

    photo_df = pd.concat([df_photod_noD, df_photoi_noD], ignore_index=True)
    n_df =  df_neutrals[ ~df_neutrals["R1"].isin(D_species) & ~df_neutrals["R2"].isin(D_species) & ~df_neutrals["R3"].isin(D_species) ]
    i_df =  df_ions[ ~df_ions["R1"].isin(D_species) & ~df_ions["R2"].isin(D_species) ]
else: 
    photo_df = pd.concat([df_photod, df_photoi], ignore_index=True)
    n_df =  copy.deepcopy(df_neutrals)
    i_df =  copy.deepcopy(df_ions)

# ===================================================================================================#

#                                    PHOTODISSOCIATION/IONIZATION

# ===================================================================================================#

# Blank out any nans and nones
photo_df[['BR']] = photo_df[['BR']].fillna("")
photo_df[['Reference']] = photo_df[['Reference']].fillna("")
photo_df['P2'] = photo_df['P2'].replace(to_replace="none", value='')
photo_df['P3'] = photo_df['P3'].replace(to_replace="none", value='')
photo_df[['P3']] = photo_df[['P3']].fillna("")

# Add reaction number column
photo_df.loc[:, "Num"] = range(1, len(photo_df)+1)
photo_df.loc[:, "Rate"] = ""
photo_df.reset_index(inplace=True)
photo_df = photo_df.drop(["index"], axis=1) # don't need this column

# Sort
if sort_by_reactants==True:
    photo_df.sort_values(["R1", "P1", "P2", "P3"], inplace=True, ignore_index=True)
else:
    # Sort by numerical column rate
    photo_df.sort_values("ColumnRate", ascending=False, inplace=True, ignore_index=True)


# Re-do the reaction number column
photo_df.loc[:, "Num"] = range(1, len(photo_df)+1)

# Assign reaction numbers
next_rxn_number = assign_reaction_numbers(photo_df)

# Make the strings 
photo_df = photo_df.assign(reactants = [f"$\mathrm{{{participant_str(r1)}}}$" if r1!="" else "" for r1 in photo_df["R1"]] )
photo_df = photo_df.assign(products = [f"$\mathrm{{{participant_str(p1,p2,p3)}}}$" for (p1,p2,p3) in zip(photo_df["P1"], photo_df["P2"], photo_df["P3"])])
photo_df = photo_df.assign(**{'Column Rate': ["$"+format(cr, '.2E').replace("-0", "-").replace("+0", "").replace("+", "").replace("E", " \times 10^{")+"}$" for cr in photo_df["ColumnRate"]]}) # colrate


# Make the final df
photodf_final = photo_df.drop(["R1", "P1", "P2", "P3", "ColumnRate"], axis=1)
photodf_final = photodf_final[["Num", "reactants", "products", "Rate", "Column Rate", "Reference"]]

# We save the dataframe into a latex table
photodf_final.style.format(escape="latex")  
photodf_final.to_latex(main_folder + "ReactionTables/photo_reactions.tex", index=False, label="photo_rxns", longtable=True, escape=False)

# ===================================================================================================#

#                                    NEUTRALS                      

# ===================================================================================================#
# Blank out any nans
for col in ["R2", "R3","P2","P3","type","M2","M1","pow","BR", "kA", "kB", "kC", "k0A","k0B","k0C","F"]:
    n_df.loc[:, col] = n_df.loc[:, col].fillna("")


if sort_by_reactants==True:
    n_df.sort_values(["R1", "R2", "R3", "P1", "P2"], inplace=True, ignore_index=True)
else:
    # Sort by numerical column rate
    n_df.sort_values("ColumnRate", ascending=False, inplace=True, ignore_index=True)

# Add reaction number column
n_df.loc[:, "Num"] = range(1, len(n_df)+1)
n_df.reset_index(inplace=True)
n_df = n_df.drop(["index"], axis=1)

# Assign reaction numbers
thenext_rxn_number = assign_reaction_numbers(n_df, opt_start=next_rxn_number)

# Make the strings 
n_df = n_df.assign(reactants = [f"$\mathrm{{{participant_str(r1,r2,r3)}}}$" if r1!="" else "" for (r1,r2,r3) in zip(n_df["R1"], n_df["R2"], n_df["R3"])])
n_df = n_df.assign(products = [f"$\mathrm{{{participant_str(p1,p2,p3)}}}$" if p1!="" else ""  for (p1,p2,p3) in zip(n_df["P1"], n_df["P2"], n_df["P3"])])

# Other column strings
n_df = n_df.assign(ratecoeff = [k_str(A,B,C,a0=A0,b0=B0,c0=C0,T="T_n") for (A,B,C,A0,B0,C0) in zip(n_df["kA"], n_df["kB"], n_df["kC"], n_df["k0A"], n_df["k0B"], n_df["k0C"])])
n_df = n_df.assign(type = [f"{typestr(t)}" for t in n_df["type"]])
n_df = n_df.assign(BR = [f"{br_str(BR)}" for BR in n_df["BR"]])
n_df = n_df.assign(MS = [f"{mscale_str(M2, M1, p)}" for (M2, M1, p) in zip(n_df["M2"], n_df["M1"], n_df["pow"])])
n_df = n_df.assign(Ref = [f"{refstr(r)}" for r in n_df["Reference"]])
n_df = n_df.assign(F = [f"{f_str(f)}" for f in n_df["F"]]) # No D reactions have troe parameter.
n_df = n_df.assign(**{'Column Rate': ["$"+format(cr, '.2E').replace("-0", "-").replace("+0", "").replace("+", "").replace("E", " \times 10^{")+"}$" for cr in n_df["ColumnRate"]]}) # colrate

# Make the final df
n_df.rename(columns = {'ratecoeff':'Rate coefficient'}, inplace = True)
ndf_final = n_df.drop(["R1", "R2", "R3", "P1", "P2", "P3", "M2", "M1", "pow",  "kA", "kB", "kC", "k0A", "k0B", "k0C", "Reference", "ColumnRate"], axis=1)
ndf_final = ndf_final[["Num", "reactants", "products", "BR", "MS", "Rate coefficient", "F", "Column Rate", "Ref"]]

# We save the dataframe into a latex table
ndf_final.style.format(escape="latex")  
ndf_final.to_latex(main_folder + "ReactionTables/neutral_reactions.tex", index=False, label="neutral_rxns", longtable=True, escape=False)

# ===================================================================================================#

#                                               IONS               

# ===================================================================================================#

for col in ["R2", "P2", "P3", "M2", "M1", "pow", "BR", "kA", "kB", "kC"]:
    i_df.loc[:, col] = i_df.loc[:, col].fillna("")

if sort_by_reactants==True:
    i_df.sort_values(["R1", "R2", "P1", "P2", "P3"], inplace=True, ignore_index=True)
else:
    # Sort by numerical column rate
    i_df.sort_values("ColumnRate", ascending=False, inplace=True, ignore_index=True)

# Add reaction number column
i_df.loc[:, "Num"] = range(1, len(i_df)+1)
i_df.reset_index(inplace=True)
i_df = i_df.drop(["index"], axis=1)

# Assign reaction numbers
assign_reaction_numbers(i_df, opt_start=thenext_rxn_number)

# Make the strings 
i_df = i_df.assign(reactants = [f"$\mathrm{{{participant_str(r1,r2)}}}$" for (r1,r2) in zip(i_df["R1"], i_df["R2"])])
i_df = i_df.assign(products = [f"$\mathrm{{{participant_str(p1,p2,p3)}}}$" for (p1,p2,p3) in zip(i_df["P1"], i_df["P2"], i_df["P3"])])

# Other column strings
i_df = i_df.assign(ratecoeff = [k_str(A,B,C) for (A,B,C) in zip(i_df["kA"], i_df["kB"], i_df["kC"])])
i_df = i_df.assign(BR = [f"{br_str(BR)}" for BR in i_df["BR"]])
i_df = i_df.assign(MS = [f"{mscale_str(M2, M1, p)}" for (M2, M1, p) in zip(i_df["M2"], i_df["M1"], i_df["pow"])])
i_df = i_df.assign(Ref = [f"{refstr(r)}" for r in i_df["Reference"]])
i_df = i_df.assign(F = [f"" for i in i_df["Reference"]]) # No D reactions have troe parameter.
i_df = i_df.assign(**{'Column Rate': [re.sub(r'E[+-]0', ' \times 10^{' + f"{'-' if '{:.2E}'.format(cr)[5] == '-' else ''}", '{:.2E}'.format(cr)+'}') for cr in i_df["ColumnRate"]]}) # colrate
i_df = i_df.assign(**{'Column Rate': ["$"+format(cr, '.2E').replace("-0", "-").replace("+0", "").replace("+", "").replace("E", " \times 10^{")+"}$" for cr in i_df["ColumnRate"]]}) # colrate


# Make the final df
i_df.rename(columns = {'ratecoeff':'Rate coefficient'}, inplace = True)
idf_final = i_df.drop(["R1", "R2", "P1", "P2", "P3", "M2", "M1", "pow",  "kA", "kB", "kC", "Reference", "ColumnRate"], axis=1)
idf_final = idf_final[["Num", "reactants", "products",  "BR", "MS", "Rate coefficient", "F",  "Column Rate", "Ref"]]

# We save the dataframe into a latex table
idf_final.style.format(escape="latex")  
idf_final.to_latex(main_folder + "ReactionTables/ion_reactions.tex", index=False, label="ion_rxns", longtable=True, escape=False)